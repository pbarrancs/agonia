from __future__ import annotations

import csv
from pathlib import Path
from typing import Dict, List, Tuple

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_XLSX_PATH = ROOT / 'inventario_agonia.xlsx'
DEFAULT_CSV_PATH = ROOT / 'new-astro-site' / 'src' / 'data' / 'productos.csv'


def normalize_inventory_workbook(xlsx_path: str | Path = DEFAULT_XLSX_PATH):
    """Normalize inventory workbook values and compute a derived Stock column.

    The workbook is expected to have the columns:
    Tipo, Color, Diseño, Talla, Unidades, Vendidas, Stock
    """
    xlsx_path = Path(xlsx_path)
    workbook = openpyxl.load_workbook(xlsx_path, data_only=True)
    sheet = workbook.active

    headers = [cell.value for cell in sheet[1]]
    if not headers:
        raise ValueError('Inventory workbook is empty')

    header_map = {name: idx + 1 for idx, name in enumerate(headers)}
    required_cols = ['Tipo', 'Color', 'Diseño',
                     'Talla', 'Unidades', 'Vendidas']
    missing = [col for col in required_cols if col not in header_map]
    if missing:
        raise ValueError(f'Missing required columns: {missing}')

    rows: List[Dict[str, object]] = []
    for row_idx in range(2, sheet.max_row + 1):
        values = [sheet.cell(row=row_idx, column=col).value for col in range(
            1, len(headers) + 1)]
        if not any(v is not None and str(v).strip() != '' for v in values):
            continue

        tipo = values[header_map['Tipo'] - 1]
        color = values[header_map['Color'] - 1]
        diseno = values[header_map['Diseño'] - 1]
        talla = values[header_map['Talla'] - 1]
        unidades = values[header_map['Unidades'] - 1]
        vendidas = values[header_map['Vendidas'] - 1]

        units = int(unidades) if unidades not in (None, '') else 0
        sold = int(vendidas) if vendidas not in (None, '') else 0
        stock = units - sold

        rows.append({
            'row': row_idx,
            'tipo': tipo,
            'color': color,
            'diseno': diseno,
            'talla': talla,
            'unidades': units,
            'vendidas': sold,
            'stock': stock,
        })

    stock_col = header_map.get('Stock', None)
    if stock_col is None:
        sheet.cell(row=1, column=len(headers) + 1, value='Stock')
        stock_col = len(headers) + 1

    for entry in rows:
        sheet.cell(
            row=entry['row'], column=header_map['Unidades']).value = entry['unidades']
        sheet.cell(
            row=entry['row'], column=header_map['Vendidas']).value = entry['vendidas']
        sheet.cell(row=entry['row'], column=stock_col).value = entry['stock']

    workbook.save(xlsx_path)
    return workbook, sheet, rows


def map_inventory_to_catalog(xlsx_path: str | Path = DEFAULT_XLSX_PATH):
    workbook, _, rows = normalize_inventory_workbook(xlsx_path)
    inventory = {}
    for entry in rows:
        key = (str(entry['tipo']), str(entry['color']),
               str(entry['diseno']), str(entry['talla']))
        inventory[key] = {
            'unidades': int(entry['unidades']),
            'vendidas': int(entry['vendidas']),
            'stock': int(entry['stock']),
        }
    return inventory


def update_catalog_availability(prod_path: str | Path = DEFAULT_CSV_PATH, inventory=None):
    prod_path = Path(prod_path)
    with prod_path.open(newline='', encoding='utf-8') as f:
        prod_rows = list(csv.DictReader(f))

    if inventory is None:
        inventory = map_inventory_to_catalog(DEFAULT_XLSX_PATH)

    prefix_to_tipo = {
        'solido': 'Solido',
        'deslavada': 'Deslavada',
        'top': 'Top',
        'sudadera': 'Sudadera',
    }
    csv_color_alias = {'Morada': 'Morado'}

    changed = []
    for pr in prod_rows:
        modelo = pr['modelo']
        parts = modelo.rsplit('_', 1)
        if len(parts) != 2:
            continue
        prefix, diseno = parts
        tipo_excel = prefix_to_tipo.get(prefix)
        if not tipo_excel:
            continue

        inv_color = csv_color_alias.get(pr['color'], pr['color'])
        key = (tipo_excel, inv_color, diseno, pr['talla'])
        entry = inventory.get(key)
        if entry is None:
            continue

        available = entry['stock']
        new_disp = 'false' if available <= 0 else 'true'
        if new_disp != pr['disponible']:
            changed.append(
                (modelo, pr['color'], pr['talla'], pr['disponible'], new_disp))
        pr['disponible'] = new_disp

    with prod_path.open('w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=prod_rows[0].keys())
        writer.writeheader()
        writer.writerows(prod_rows)

    return prod_rows, changed
