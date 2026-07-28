#!/usr/bin/env python3
"""
create_excel.py — rebuild inventario_agonia.xlsx from the markdown inventory table.

Disclaimer:
This helper is meant to reconstruct the workbook from the documented inventory source.
It is not a destructive inventory editor. By default it will NOT overwrite an existing
inventario_agonia.xlsx; it exits safely unless you pass --force.

Purpose:
- Read the inventory table from Documentos/playeras/detalles_agonia_playeras.md
- Generate or refresh inventario_agonia.xlsx at the repository root
- Keep the workbook structure aligned with the current catalog workflow
"""
import argparse
import re
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

REPO_ROOT = Path(__file__).resolve().parent.parent
MD_PATH = REPO_ROOT / 'Documentos' / 'playeras' / 'detalles_agonia_playeras.md'
OUTPUT_PATH = REPO_ROOT / 'inventario_agonia.xlsx'


def parse_inventory_rows(md_path: Path):
    if not md_path.exists():
        sys.exit(f'ERROR: inventory markdown not found: {md_path}')

    with md_path.open(encoding='utf-8') as handle:
        content = handle.read()

    start = content.find('## Inventario Tabla')
    if start == -1:
        sys.exit('ERROR: "## Inventario Tabla" not found')

    lines = content[start:].splitlines()
    rows = []
    state = 'seek_header'

    for line in lines:
        stripped = line.strip()
        if state == 'seek_header':
            if stripped.startswith('|') and 'Tipo' in stripped and 'Color' in stripped:
                state = 'seek_sep'
        elif state == 'seek_sep':
            if re.match(r'^\|[\s\-|]+\|$', stripped):
                state = 'data'
        elif state == 'data':
            if not stripped.startswith('|'):
                break
            # Keep positional indices — blank cells must NOT be filtered out.
            # Row format: | col1 | col2 | col3 | col4 | col5 | col6 |
            # After split('|'): ['', col1, col2, col3, col4, col5, col6, '']
            parts = [p.strip() for p in stripped.split('|')]
            if len(parts) < 7:  # need indices 1-6
                continue
            tipo, color, diseno, unidades_s, vendidas_s, talla = (
                parts[1], parts[2], parts[3], parts[4], parts[5], parts[6]
            )
            unidades = int(unidades_s) if unidades_s.isdigit() else 0
            vendidas = int(vendidas_s) if vendidas_s.isdigit() else 0
            rows.append((tipo, color, diseno, talla, unidades, vendidas))

    return rows


def build_workbook(rows, output_path: Path, force: bool = False):
    if output_path.exists() and not force:
        print(f'Workbook already exists at: {output_path}')
        print('No changes were made. Re-run with --force if you intentionally want to overwrite it.')
        return False

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Inventario'

    headers = ['Tipo', 'Color', 'Diseño',
               'Talla', 'Unidades', 'Vendidas', 'Stock']
    h_font = Font(bold=True, color='FFFFFF', name='Calibri', size=11)
    h_fill = PatternFill('solid', fgColor='1A1A1A')
    h_align = Alignment(horizontal='center', vertical='center')

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = h_font
        cell.fill = h_fill
        cell.alignment = h_align

    ws.row_dimensions[1].height = 22

    tipo_fill = {
        'Solido': 'F5F5F5',
        'Deslavada': 'E8F5E9',
        'Top': 'E3F2FD',
        'Manga Larga': 'FFF3E0',
        'Sudadera': 'FCE4EC',
    }

    for row_idx, (tipo, color, diseno, talla, unidades, vendidas) in enumerate(rows, 2):
        fill_color = tipo_fill.get(tipo, 'FFFFFF')
        fill = PatternFill('solid', fgColor=fill_color)
        stock = unidades - vendidas
        for col, value in enumerate([tipo, color, diseno, talla, unidades, vendidas, stock], 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.fill = fill
            cell.alignment = Alignment(horizontal='center')

    for letter, width in zip('ABCDEFG', [14, 12, 8, 8, 12, 12, 10]):
        ws.column_dimensions[letter].width = width

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:G{len(rows) + 1}'

    wb.save(output_path)
    print(f'Saved: {output_path}')
    print(f'Total rows: {len(rows)}')
    return True


if __name__ == '__main__':
    parser = argparse.ArgumentParser(
        description='Rebuild inventario_agonia.xlsx from the markdown inventory table.')
    parser.add_argument('--force', action='store_true',
                        help='overwrite an existing inventario_agonia.xlsx')
    args = parser.parse_args()

    rows = parse_inventory_rows(MD_PATH)
    print(f'Parsed {len(rows)} rows from inventory table.')
    build_workbook(rows, OUTPUT_PATH, force=args.force)
