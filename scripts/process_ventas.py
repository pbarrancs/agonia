import csv
import openpyxl
from collections import defaultdict, Counter

csv_path = r'c:\Users\pablo\OneDrive\Documentos\Docs\agonia\ventas.csv'
xlsx_path = r'c:\Users\pablo\OneDrive\Documentos\Docs\agonia\inventario_agonia.xlsx'
prod_path = r'c:\Users\pablo\OneDrive\Documentos\Docs\agonia\new-astro-site\src\data\productos.csv'

# ─── Read ventas ──────────────────────────────────────────────────────────────
with open(csv_path, newline='', encoding='utf-8') as f:
    ventas = list(csv.DictReader(f))

DESLAVADA_COLORS = {'Morado', 'Gris'}


def infer_tipo_excel(row):
    color = row['color']
    tipo_str = row['tipo']
    produto = row['producto']

    if not color or not tipo_str:
        return None

    diseno = tipo_str.replace('Tipo ', '')

    if produto == 'Sudadera':
        tipo_excel = 'Sudadera'
    elif produto == 'Tank Top':
        tipo_excel = 'Top'
    elif color in DESLAVADA_COLORS:
        tipo_excel = 'Deslavada'
    else:
        tipo_excel = 'Solido'

    return (tipo_excel, diseno)


# ─── Load inventory ───────────────────────────────────────────────────────────
wb = openpyxl.load_workbook(xlsx_path)
ws = wb.active

inv = {}
for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
    tipo, color, diseno, talla, unidades, vendidas = row
    key = (str(tipo), str(color), str(diseno), str(talla))
    inv[key] = {'unidades': int(unidades or 0),
                'vendidas': int(vendidas or 0), 'row': idx}

# ─── Match sales to inventory ─────────────────────────────────────────────────
TALLA_NORM = {
    'XC': 'XC', 'XS': 'XS', 'S': 'S', 'M': 'M',
    'G': 'G', 'L': 'G', 'XL': 'XL',
    'S/M': None, 'M/L': None, '': ''
}

unmatched = []
matched_sales = []

for venta in ventas:
    result = infer_tipo_excel(venta)
    if result is None:
        unmatched.append((venta['nombre'], venta['producto'], venta['color'],
                          venta['talla'], venta['tipo'], 'missing color or tipo'))
        continue

    tipo_excel, diseno = result
    talla_raw = venta['talla']
    talla = TALLA_NORM.get(talla_raw)

    if talla is None:
        unmatched.append((venta['nombre'], venta['producto'], venta['color'],
                          venta['talla'], venta['tipo'], 'talla S/M or M/L not in inventory'))
        continue
    if talla == '':
        unmatched.append((venta['nombre'], venta['producto'], venta['color'],
                          venta['talla'], venta['tipo'], 'no talla'))
        continue

    key = (tipo_excel, venta['color'], diseno, talla)
    if key not in inv:
        unmatched.append((venta['nombre'], venta['producto'], venta['color'],
                          venta['talla'], venta['tipo'], f'key not found: {key}'))
        continue

    matched_sales.append((key, venta['nombre']))

print(f'Matched: {len(matched_sales)}, Unmatched: {len(unmatched)}')

# ─── Apply to inventory ───────────────────────────────────────────────────────
sales_count = defaultdict(int)
for (key, nombre) in matched_sales:
    sales_count[key] += 1

negative_flags = []
for key, count in sales_count.items():
    entry = inv[key]
    new_vendidas = entry['vendidas'] + count
    available = entry['unidades'] - new_vendidas
    row_idx = entry['row']
    ws.cell(row=row_idx, column=6).value = new_vendidas
    if available < 0:
        negative_flags.append(
            (key, entry['unidades'], new_vendidas, available))
        comment = openpyxl.comments.Comment(
            'stock negativo — revisar', 'agonia')
        ws.cell(row=row_idx, column=6).comment = comment
    entry['vendidas'] = new_vendidas

wb.save(xlsx_path)
print('inventario_agonia.xlsx updated')

# ─── Update productos.csv ─────────────────────────────────────────────────────
prefix_to_tipo = {
    'solido': 'Solido', 'deslavada': 'Deslavada',
    'top': 'Top', 'sudadera': 'Sudadera'
}

with open(prod_path, newline='', encoding='utf-8') as f:
    prod_rows = list(csv.DictReader(f))

changed_oos = []
for pr in prod_rows:
    modelo = pr['modelo']
    parts = modelo.rsplit('_', 1)
    if len(parts) != 2:
        continue
    prefix, diseno = parts
    tipo_excel = prefix_to_tipo.get(prefix)
    if not tipo_excel:
        continue

    key = (tipo_excel, pr['color'], diseno, pr['talla'])
    if key in inv:
        entry = inv[key]
        available = entry['unidades'] - entry['vendidas']
        new_disp = 'false' if available <= 0 else 'true'
        if new_disp != pr['disponible']:
            changed_oos.append(
                (modelo, pr['color'], pr['talla'], pr['disponible'], new_disp))
        pr['disponible'] = new_disp

with open(prod_path, 'w', newline='', encoding='utf-8') as f:
    w = csv.DictWriter(f, fieldnames=prod_rows[0].keys())
    w.writeheader()
    w.writerows(prod_rows)
print(f'productos.csv updated. {len(changed_oos)} availability changes')

# ─── REPORT ───────────────────────────────────────────────────────────────────
print()
print('=' * 64)
print('REPORT')
print('=' * 64)

print()
print('1. UNITS SOLD PER TIPO / COLOR (matched only)')
by_tipo_color = Counter()
for (key, nombre) in matched_sales:
    tipo_e, color, diseno, talla = key
    by_tipo_color[('Tipo ' + diseno, color, tipo_e)] += 1
for (tipo, color, subtipo), cnt in sorted(by_tipo_color.items()):
    print(f'  {tipo} | {subtipo:10} | {color:8}: {cnt}')

total_matched = len(matched_sales)
print(f'  TOTAL matched: {total_matched}')

print()
print('2. ROWS FLAGGED revisar=Si')
with open(csv_path, newline='', encoding='utf-8') as f:
    for r in csv.DictReader(f):
        if r['revisar'] == 'Si':
            print(
                f"  {r['nombre']:22} | {r['producto']:10} | color={r['color']:8} talla={r['talla']:4} tipo={r['tipo']:6} | {r['gym']}")

print()
print('3. UNMATCHED (inventory not updated for these)')
for item in unmatched:
    print(
        f'  {item[0]:22} | {item[1]:10} | color={item[2]:8} talla={item[3]:4} tipo={item[4]:6} | {item[5]}')

print()
print('4. STOCK CHANGES IN productos.csv')
for (modelo, color, talla, old, new) in changed_oos:
    arrow = 'NOW OOS' if new == 'false' else 'NOW AVAILABLE'
    print(f'  {modelo:16} | {color:8} | {talla:4} -> {arrow} (was {old})')

if negative_flags:
    print()
    print('!! NEGATIVE STOCK FLAGS')
    for (key, unid, vend, avail) in negative_flags:
        print(f'  {key} unidades={unid} vendidas={vend} net={avail}')
else:
    print()
    print('No negative stock detected.')
