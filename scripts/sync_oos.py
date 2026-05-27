"""
sync_oos.py — Sync all zero/negative stock rows to:
  1. productos.csv  -> disponible = false
  2. inventario_agonia.xlsx -> comment on rows with negative stock
"""
import csv
import openpyxl

xlsx_path = r'c:\Users\pablo\OneDrive\Documentos\Docs\agonia\inventario_agonia.xlsx'
prod_path = r'c:\Users\pablo\OneDrive\Documentos\Docs\agonia\new-astro-site\src\data\productos.csv'

# ─── Load inventory ───────────────────────────────────────────────────────────
wb = openpyxl.load_workbook(xlsx_path)
ws = wb.active

inv = {}  # (tipo, color, diseno, talla) -> {unidades, vendidas, row}
for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
    tipo, color, diseno, talla, unidades, vendidas = row
    key = (str(tipo), str(color), str(diseno), str(talla))
    inv[key] = {
        'unidades': int(unidades or 0),
        'vendidas': int(vendidas or 0),
        'row': idx,
    }

# ─── Add/clear Excel comments based on availability ──────────────────────────
neg_count = 0
for key, entry in inv.items():
    available = entry['unidades'] - entry['vendidas']
    cell = ws.cell(row=entry['row'], column=6)
    if available < 0:
        if cell.comment is None:
            cell.comment = openpyxl.comments.Comment(
                'stock negativo — revisar', 'agonia')
        neg_count += 1

wb.save(xlsx_path)
print(f'Excel updated: {neg_count} negative-stock rows flagged')

# ─── Update productos.csv ─────────────────────────────────────────────────────
prefix_to_tipo = {
    'solido':    'Solido',
    'deslavada': 'Deslavada',
    'top':       'Top',
    'sudadera':  'Sudadera',
}

# CSV uses Spanish gender variants; inventory uses a single canonical form
CSV_COLOR_ALIAS = {
    'Morada': 'Morado',
}

with open(prod_path, newline='', encoding='utf-8') as f:
    prod_rows = list(csv.DictReader(f))

now_oos = []
now_avail = []

for pr in prod_rows:
    modelo = pr['modelo']
    parts = modelo.rsplit('_', 1)
    if len(parts) != 2:
        continue
    prefix, diseno = parts
    tipo_excel = prefix_to_tipo.get(prefix)
    if not tipo_excel:
        continue

    csv_color = pr['color']
    inv_color = CSV_COLOR_ALIAS.get(csv_color, csv_color)
    key = (tipo_excel, inv_color, diseno, pr['talla'])
    if key not in inv:
        continue

    entry = inv[key]
    available = entry['unidades'] - entry['vendidas']
    new_disp = 'false' if available <= 0 else 'true'

    if new_disp != pr['disponible']:
        if new_disp == 'false':
            now_oos.append((modelo, pr['color'], pr['talla']))
        else:
            now_avail.append((modelo, pr['color'], pr['talla']))
    pr['disponible'] = new_disp

with open(prod_path, 'w', newline='', encoding='utf-8') as f:
    w = csv.DictWriter(f, fieldnames=prod_rows[0].keys())
    w.writeheader()
    w.writerows(prod_rows)

print(
    f'productos.csv updated: {len(now_oos)} newly OOS, {len(now_avail)} newly available')

if now_oos:
    print('\nNow OOS:')
    for m, c, t in sorted(now_oos):
        print(f'  {m:16} | {c:8} | {t}')

if now_avail:
    print('\nNow available (were OOS):')
    for m, c, t in sorted(now_avail):
        print(f'  {m:16} | {c:8} | {t}')

# Summary: total OOS in catalog after update
total_oos = sum(1 for pr in prod_rows if pr['disponible'] == 'false')
total = len(prod_rows)
print(
    f'\nTotal catalog rows: {total} | OOS: {total_oos} | Available: {total - total_oos}')
