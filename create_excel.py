#!/usr/bin/env python3
"""
create_excel.py  —  generates inventario_agonia.xlsx at the repo root.
Reads the inventory table from Documentos/playeras/detalles_agonia_playeras.md.
"""
import os
import re
import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

ROOT = os.path.dirname(os.path.abspath(__file__))
MD = os.path.join(ROOT, 'Documentos', 'playeras',
                  'detalles_agonia_playeras.md')

with open(MD, encoding='utf-8') as f:
    content = f.read()

# Locate "## Inventario Tabla"
start = content.find('## Inventario Tabla')
if start == -1:
    sys.exit('ERROR: "## Inventario Tabla" not found')

lines = content[start:].splitlines()

# State machine: seek header → seek separator → read data
rows = []
state = 'seek_header'

for line in lines:
    stripped = line.strip()
    if state == 'seek_header':
        if stripped.startswith('|') and 'Tipo' in stripped and 'Color' in stripped:
            state = 'seek_sep'
    elif state == 'seek_sep':
        if re.match(r'^\|[\s\-|]+\|$', stripped):
            state = 'data'
    elif state == 'data':
        if not stripped.startswith('|'):
            break
        # Keep positional indices — blank cells must NOT be filtered out.
        # Row format: | col1 | col2 | col3 | col4 | col5 | col6 |
        # After split('|'): ['', col1, col2, col3, col4, col5, col6, '']
        parts = [p.strip() for p in stripped.split('|')]
        if len(parts) < 7:          # need indices 1-6
            continue
        tipo, color, diseno, unidades_s, vendidas_s, talla = (
            parts[1], parts[2], parts[3], parts[4], parts[5], parts[6]
        )
        unidades = int(unidades_s) if unidades_s.isdigit() else 0
        vendidas = int(vendidas_s) if vendidas_s.isdigit() else 0
        rows.append((tipo, color, diseno, talla, unidades, vendidas))

print(f'Parsed {len(rows)} rows from inventory table.')

# ── Build workbook ───────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Inventario'

HEADERS = ['Tipo', 'Color', 'Diseño', 'Talla', 'Unidades', 'Vendidas']
H_FONT = Font(bold=True, color='FFFFFF', name='Calibri', size=11)
H_FILL = PatternFill('solid', fgColor='1A1A1A')
H_ALIGN = Alignment(horizontal='center', vertical='center')

for col, h in enumerate(HEADERS, 1):
    c = ws.cell(row=1, column=col, value=h)
    c.font = H_FONT
    c.fill = H_FILL
    c.alignment = H_ALIGN

ws.row_dimensions[1].height = 22

TIPO_FILL = {
    'Solido':      'F5F5F5',
    'Deslavada':   'E8F5E9',
    'Top':         'E3F2FD',
    'Manga Larga': 'FFF3E0',
    'Sudadera':    'FCE4EC',
}

for r, (tipo, color, diseno, talla, unidades, vendidas) in enumerate(rows, 2):
    fc = TIPO_FILL.get(tipo, 'FFFFFF')
    fill = PatternFill('solid', fgColor=fc)
    for col, val in enumerate([tipo, color, diseno, talla, unidades, vendidas], 1):
        c = ws.cell(row=r, column=col, value=val)
        c.fill = fill
        c.alignment = Alignment(horizontal='center')

# Column widths
for letter, width in zip('ABCDEF', [14, 12, 8, 8, 12, 12]):
    ws.column_dimensions[letter].width = width

ws.freeze_panes = 'A2'
ws.auto_filter.ref = f'A1:F{len(rows) + 1}'

out = os.path.join(ROOT, 'inventario_agonia.xlsx')
wb.save(out)
print(f'Saved: {out}')
print(f'Total rows: {len(rows)}')
